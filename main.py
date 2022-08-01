import openpyxl as opx;
tabela = opx.load_workbook(filename='relatorios.xlsx')
tabelaantiga = tabela['RelatorioRequisicaoPaciente']
for i in range(3, 464):
    val = tabelaantiga.cell(row=i, column=2).value.split()
    tabelaantiga.cell(row=i, column=2).value = val[0]

tabela.save(filename='relatorios.xlsx')